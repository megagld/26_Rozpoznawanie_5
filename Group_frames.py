import os
from openpyxl import load_workbook

def group(file_to_group):
   
    path = os.getcwd()
    xlsx_file='{}/data/times.xlsx'.format(path)

    wb=load_workbook(xlsx_file)
    ws = wb['times']
 
    detected_times=[]

    for i in ws.values:

        if i[0]!='file' and i[0]==file_to_group:
            detected_times.append(i[1])

    wb.close()

    # setting max gap
    max_gap=1000*3 # ~ 3sec

    grouped_times=[]
    tmp=[]
    while detected_times:
        
        frame_to_analize=detected_times.pop(0)
        if tmp and frame_to_analize-tmp[-1]>max_gap:
            grouped_times.append(tmp)
            tmp=[]          
        tmp.append(frame_to_analize)

    grouped_times.append(tmp)

    # making a start and end cuts

    if  grouped_times==[[]]:
        print(f"Plik {file_to_group} -nie wykryto klatek z rowerem/osobą")
    else:
        video_cuts=[(min(i),max(i)) for i in grouped_times]

        # saving cuts to xslx

        path = os.getcwd()
        xlsx_file='{}/data/times.xlsx'.format(path)
        wb=load_workbook(xlsx_file)
        ws = wb['cuts']
        last_empty=ws.max_row


        for i,j in enumerate(video_cuts):

            ws.cell(row=last_empty+i+1,column=1).value=file_to_group
            ws.cell(row=last_empty+i+1,column=2).value=i        
            ws.cell(row=last_empty+i+1,column=3).value=j[0]
            ws.cell(row=last_empty+i+1,column=4).value=j[1]

        wb.save(xlsx_file)
        wb.close()
        print(f"Plik {file_to_group} -klatki zostały pogrupowane")
    
def main():
    for i,j,k in os.walk('data'):
        for l in k:
            if l[-4:]=='.mp4':
                group(l)

if __name__ == "__main__":
    main()
