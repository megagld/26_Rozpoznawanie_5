import cv2
from ultralytics import YOLO
import math
from openpyxl import load_workbook, Workbook
import os
import ffmpeg



def detect_bike(file,path):

    # self.name_label.configure(text=file)


    input_video='{}/{}'.format(path,file)

    # input video
    cap = cv2.VideoCapture(input_video)


    #total time
    # uses ffprobe command to extract all possible metadata from the media file
    total_time=float(ffmpeg.probe(input_video)["streams"][0]['duration'])
        

    # model
    model = YOLO("yolo-Weights/yolov8n.pt")

    # object classes
    classNames = ["person", "bicycle", "car", "motorbike", "aeroplane", "bus", "train", "truck", "boat",
                "traffic light", "fire hydrant", "stop sign", "parking meter", "bench", "bird", "cat",
                "dog", "horse", "sheep", "cow", "elephant", "bear", "zebra", "giraffe", "backpack", "umbrella",
                "handbag", "tie", "suitcase", "frisbee", "skis", "snowboard", "sports ball", "kite", "baseball bat",
                "baseball glove", "skateboard", "surfboard", "tennis racket", "bottle", "wine glass", "cup",
                "fork", "knife", "spoon", "bowl", "banana", "apple", "sandwich", "orange", "broccoli",
                "carrot", "hot dog", "pizza", "donut", "cake", "chair", "sofa", "pottedplant", "bed",
                "diningtable", "toilet", "tvmonitor", "laptop", "mouse", "remote", "keyboard", "cell phone",
                "microwave", "oven", "toaster", "sink", "refrigerator", "book", "clock", "vase", "scissors",
                "teddy bear", "hair drier", "toothbrush"
                ]

    detected_bikes_times=[]
    detected_bikes_times_xlsx=[]
    time_count=0
    time_detect_freq=1000

    while True:

        # frame_count+=frames_detect_freq
        time_count+=time_detect_freq
        
        if time_count>1*total_time:
            break

        cap.set(0,time_count)
        res, img = cap.read()

        results = model.predict(source=img, classes=[0,1])

        print(time_count)
        # coordinates
        for r in results:
            boxes = r.boxes

            for box in boxes:
                # class name
                cls = int(box.cls[0])

                # bounding box
                x1, y1, x2, y2 = box.xyxy[0]
                x1, y1, x2, y2 = int(x1), int(y1), int(x2), int(y2) # convert to int values

                # confidence
                confidence = math.ceil((box.conf[0]*100))/100

                #found_rec
                detected_bikes_times.append(time_count)

                detected_object=classNames[cls]

                # save found_rec to xlsx

                detected_bikes_times_xlsx.append((file,time_count,x1,y1,x2,y2,confidence,detected_object))

        if img is None:
            break

    # removing duplicated frames
    detected_bikes_times=sorted(set(detected_bikes_times))

    # saving detected_frames to xlsx

    xlsx_file='{}/test.xlsx'.format(path)
    wb=load_workbook(xlsx_file)
    ws = wb['times']
    last_empty=ws.max_row
    for i,j in enumerate(detected_bikes_times_xlsx):
        file,time_count,x1,y1,x2,y2,confidence,detected_object=j
        ws.cell(row=last_empty+1+i,column=1).value=file
        ws.cell(row=last_empty+1+i,column=2).value=time_count
        ws.cell(row=last_empty+1+i,column=3).value=x1
        ws.cell(row=last_empty+1+i,column=4).value=y1
        ws.cell(row=last_empty+1+i,column=5).value=x2
        ws.cell(row=last_empty+1+i,column=6).value=y2
        ws.cell(row=last_empty+1+i,column=7).value=confidence
        ws.cell(row=last_empty+1+i,column=8).value=detected_object

    wb.save(xlsx_file)
    wb.close()

    cap.release()
    cv2.destroyAllWindows()

file = 'test.py'
path = os.getcwd()


detect_bike(file,path)