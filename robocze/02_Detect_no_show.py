from ultralytics import YOLO
import cv2
import json
import os
from wakepy import keep
import time
from Time_counter import *
from tkinter import ttk
import winsound
import tkinter as tk
import math
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment

class App(tk.Tk):
    def __init__(self):
        super().__init__()

        # root window
        row_c=8
        row_height=25
        cols=[150,250]
        prop=cols[0]//cols[1]

        root_width=sum(cols)
        root_height=row_c*row_height

        # root = tk.Tk()
        self.geometry('{}x{}'.format(root_width,root_height))
        self.title('Rozpoznaj rowery')
        self.resizable(0, 0)
         
        # configure the grid
        self.columnconfigure(0, weight=prop)
        self.columnconfigure(1, weight=1)

        self.create_widgets()
        
    def create_widgets(self):

        #set labels
        texts=['',
            '',
            '',
            'Plik',
            '',
            'Frames',
            '',
            '']
        self.texts_state={}

        for i,j in enumerate(texts):
            label = ttk.Label(self, text=j,font=('helvetica', 10))
            label.grid(column=0, row=i, sticky=tk.E, padx=5)
            self.texts_state[i]=label

        # set entry
        entrys=['',                
                'button',
                '',
                'label_1',
                '',
                'label_2',
                '',
                'progresbar']
        self.entry_state={}

        for i,j in enumerate(entrys):
            if j=='':
                entry = ttk.Label(self, text=j,font=('helvetica', 10))
                entry.grid(column=0, row=i, sticky=tk.E, padx=5)
            elif j=='checkbox':
                self.var = tk.IntVar(value=1)
                entry=tk.Checkbutton(self,variable=self.var)
            elif j=='button':
                entry=tk.Button(text='Rozpoznaj rowery', command=self.run, bg='brown', fg='white', font=('helvetica', 10, 'bold'),width=16)
            elif j=='progresbar':
                entry=ttk.Progressbar(self, orient='horizontal',mode='determinate', length=250)
            elif j=='label_1':
                entry=ttk.Label(self, text='')
            elif j=='label_2':
                entry=ttk.Label(self, text='')
            else:
                entry = ttk.Entry(self,textvariable=j,width=30)
                entry.insert(-1, j)

            entry.grid(column=1, row=i, sticky=tk.W, padx=5)
            self.entry_state[i]=entry

        self.progressbar=self.entry_state[7]
        self.frames_label=self.entry_state[5]
        self.name_label=self.entry_state[3]

        self.act_labs=[self.progressbar,
                       self.frames_label,
                       self.name_label]


    def detect_bike(self,file,path,det_file_name):

        self.name_label.configure(text=file)

        with keep.running() as k:
            # do stuff that takes long time

            start_time = time.time()

            input_video='{}/{}'.format(path,file)

            # input video
            cap = cv2.VideoCapture(input_video)

            #frames
            self.frames=int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
            
            # model
            model = YOLO("yolo-Weights/yolov8n.pt")

            # object classes
            classNames = ["person", "bicycle", "car", "motorbike", "aeroplane", "bus", "train", "truck", "boat",
                        "traffic light", "fire hydrant", "stop sign", "parking meter", "bench", "bird", "cat",
                        "dog", "horse", "sheep", "cow", "elephant", "bear", "zebra", "giraffe", "backpack", "umbrella",
                        "handbag", "tie", "suitcase", "frisbee", "skis", "snowboard", "sports ball", "kite", "baseball bat",
                        "baseball glove", "skateboard", "surfboard", "tennis racket", "bottle", "wine glass", "cup",
                        "fork", "knife", "spoon", "bowl", "banana", "apple", "sandwich", "orange", "broccoli",
                        "carrot", "hot dog", "pizza", "donut", "cake", "chair", "sofa", "pottedplant", "bed",
                        "diningtable", "toilet", "tvmonitor", "laptop", "mouse", "remote", "keyboard", "cell phone",
                        "microwave", "oven", "toaster", "sink", "refrigerator", "book", "clock", "vase", "scissors",
                        "teddy bear", "hair drier", "toothbrush"
                        ]

            detected_frames=[]
            detected_frames_xlsx=[]
            frame_count=0
            frames_detect_freq=60

            while True:

                frame_count+=frames_detect_freq

                self.frames_label.configure(text=f'{frame_count}/{self.frames}')
                self.progressbar['value']=int(100*(frame_count/self.frames)) 
                self.update_status()

                cap.set(1,frame_count)
                res, img = cap.read()

                results = model.predict(source=img, classes=[0,1])

                print(frame_count)
                # coordinates
                for r in results:
                    boxes = r.boxes

                    for box in boxes:
                        # class name
                        cls = int(box.cls[0])

                        # bounding box
                        x1, y1, x2, y2 = box.xyxy[0]
                        x1, y1, x2, y2 = int(x1), int(y1), int(x2), int(y2) # convert to int values

                        # confidence
                        confidence = math.ceil((box.conf[0]*100))/100

                        #found_rec
                        detected_frames.append(frame_count)

                        # save found_rec to xlsx

                        detected_frames_xlsx.append((file,frame_count,x1,y1,x2,y2,confidence))
                    # try:
                    #     cv2.imshow("input", img)
                    # except:
                    #     pass

                if img is None:
                    break

                if cv2.waitKey(1) == ord('q'):
                    break

            # removing duplicated frames
            detected_frames=sorted(set(detected_frames))

            # saving detected_frames to json

            with open('{}/{}'.format(path,det_file_name), 'w') as f:
                json.dump(detected_frames, f)


            # saving detected_frames to xlsx

            xlsx_file='{}/frames.xlsx'.format(path)
            wb=load_workbook(xlsx_file)
            ws = wb['frames']
            last_empty=ws.max_row
            for i,j in enumerate(detected_frames_xlsx):
                file,frame_count,x1,y1,x2,y2,confidence=j
                ws.cell(row=last_empty+1+i,column=1).value=file
                ws.cell(row=last_empty+1+i,column=2).value=frame_count
                ws.cell(row=last_empty+1+i,column=3).value=x1
                ws.cell(row=last_empty+1+i,column=4).value=y1
                ws.cell(row=last_empty+1+i,column=5).value=x2
                ws.cell(row=last_empty+1+i,column=6).value=y2
                ws.cell(row=last_empty+1+i,column=7).value=confidence

            wb.save(xlsx_file)
            wb.close()

            cap.release()
            cv2.destroyAllWindows()

            save_time(start_time, os.path.basename(__file__).rstrip('.py'),input_video)

    def run(self):
        for path,_,files in os.walk('data'):
            for file in files:
                # det_file_name='{}'.format(file.replace('_resized.mp4','_detected_frames.json'))
                det_file_name='{}'.format(file.replace('.mp4','_detected_frames.json'))

                # if all((file.endswith('_resized.mp4'), det_file_name not in files)):
                if det_file_name not in files:

                    self.detect_bike(file,path,det_file_name)

    def update_status(self):
        for i in self.act_labs:
            i.update_idletasks()

if __name__ == "__main__":
    app = App()
    app.mainloop()