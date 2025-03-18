import json
import os
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment

def group(file_to_group):

    # getting data
    with open(file_to_group) as json_file:
        detected_frames = json.load(json_file)

    # setting frames max gap
    max_gap=240 # ~ 1sec with 24fps


    tmp=[]
    grouped_frames=[]
    for i,j in enumerate(detected_frames[:-1]):
        if detected_frames[i+1]-j<max_gap:
            tmp.append(j)
        else:
            if tmp:
                grouped_frames.append(tmp)
            tmp=[]
    if tmp:
        grouped_frames.append(tmp)

    # making a start and end cuts

    video_cuts=[(min(i),max(i)) for i in grouped_frames]

    # saving cuts to json

    with open('{}'.format(file_to_group.replace('detected_frames','video_cuts')), 'w') as f:
        json.dump(video_cuts, f)

    # saving cuts to xslx

    path = os.getcwd()
    xlsx_file='{}/data/frames.xlsx'.format(path)
    wb=load_workbook(xlsx_file)
    ws = wb['cuts']
    last_empty=ws.max_row


    for i,j in enumerate(video_cuts):

        ws.cell(row=last_empty+i+1,column=1).value=file_to_group.rstrip('_detected_frames.json').lstrip('data\\')
        ws.cell(row=last_empty+i+1,column=2).value=i        
        ws.cell(row=last_empty+i+1,column=3).value=j[0]
        ws.cell(row=last_empty+i+1,column=4).value=j[1]

    wb.save(xlsx_file)
    wb.close()

    
for i,j,k in os.walk('data'):
    for l in k:
        if 'detected_frames' in l:
            group('{}\\{}'.format(i,l))
