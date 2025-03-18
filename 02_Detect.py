from ultralytics import YOLO
import cv2
import os
from wakepy import keep
import time
from Time_counter import *
from tkinter import ttk
import winsound
import tkinter as tk
import math
from openpyxl import load_workbook
import ffmpeg
import shutil
import Group_frames
import Cut_video


class App(tk.Tk):
    def __init__(self):
        super().__init__()

        # files to detect
        self.files_to_detect = []
        self.parse_data_folder()

        # config

        window_style = ttk.Style()
        window_style.configure('wiersze', rowheight=55)

        # root window
        row_c=11+len(self.files_to_detect)
        row_height=25
        cols=[150,250]
        prop=cols[0]//cols[1]

        root_width=sum(cols)
        root_height=row_c*row_height

        # root = tk.Tk()
        self.geometry('{}x{}'.format(root_width,root_height))
        self.title('Detect bikes')
        self.resizable(0, 0)
         
        # configure the grid
        self.columnconfigure(0, weight=prop)
        self.columnconfigure(1, weight=1)

        self.create_widgets()
        
    def create_widgets(self):

        #set labels
        texts=['',
            '',
            '',
            '',
            '',
            '',
            '',
            'File',
            '',
            'Times',
            '',
            '',
            '']
        self.texts_state={}

        for i,j in enumerate(texts):
            label = ttk.Label(self, text=j,font=('helvetica', 10))
            label.grid(column=0, row=i, sticky=tk.E, padx=5)
            self.texts_state[i]=label
            
        # set entry
        entrys=['',                
                'button_1',
                '',
                'button_2',
                '',
                'button_3',
                '',
                'label_1',
                '',
                'label_2',
                '',
                'progresbar',
                '']
        self.entry_state={}

        for i,j in enumerate(entrys):
            if j=='':
                entry = ttk.Label(self, text=j,font=('helvetica', 10))
                entry.grid(column=0, row=i, sticky=tk.E, padx=5)
            elif j=='checkbox':
                self.var = tk.IntVar(value=1)
                entry=tk.Checkbutton(self,variable=self.var)
            elif j=='button_1':
                entry=tk.Button(text='Detect bikes', command=self.detect, bg='brown', fg='white', font=('helvetica', 10, 'bold'),width=16)
            elif j=='button_2':
                entry=tk.Button(text='Group frames', command=self.group, bg='brown', fg='white', font=('helvetica', 10, 'bold'),width=16)
            elif j=='button_3':
                entry=tk.Button(text='Cut video', command=self.cut, bg='brown', fg='white', font=('helvetica', 10, 'bold'),width=16)
            elif j=='progresbar':
                entry=ttk.Progressbar(self, orient='horizontal',mode='determinate', length=250)
            elif j=='label_1':
                entry=ttk.Label(self, text='')
            elif j=='label_2':
                entry=ttk.Label(self, text='')
            else:
                entry = ttk.Entry(self,textvariable=j,width=30)
                entry.insert(-1, j)

            entry.grid(column=1, row=i, sticky=tk.W, padx=5)
            self.entry_state[i]=entry

        self.progressbar=self.entry_state[11]
        self.times_label=self.entry_state[9]
        self.name_label=self.entry_state[7]

        self.act_labs=[self.progressbar,
                       self.times_label,
                       self.name_label]
    
        self.base_raw_count = i    
        # files list        
        for cnt,file,_,work_time in self.files_to_detect:
            i+=1
            entry = ttk.Label(self,text=f'{cnt} : {file} / {work_time}')
            entry.grid(column=1, row=i, sticky=tk.W, padx=5)

            self.entry_state[i] = entry
            self.act_labs.append(self.entry_state[i])

    def detect_bike(self,cnt,file,path):

        self.name_label.configure(text=file)

        with keep.running() as k:
            # do stuff that takes long time

            start_time = time.time()

            input_video='{}/{}'.format(path,file)

            # input video
            cap = cv2.VideoCapture(input_video)

            #total time
            # uses ffprobe command to extract all possible metadata from the media file
            self.total_time=float(ffmpeg.probe(input_video)["streams"][0]['duration'])
            
            # model
            model = YOLO("yolo-Weights/yolov8n.pt")

            # object classes
            classNames = ["person", "bicycle", "car", "motorbike", "aeroplane", "bus", "train", "truck", "boat",
                        "traffic light", "fire hydrant", "stop sign", "parking meter", "bench", "bird", "cat",
                        "dog", "horse", "sheep", "cow", "elephant", "bear", "zebra", "giraffe", "backpack", "umbrella",
                        "handbag", "tie", "suitcase", "frisbee", "skis", "snowboard", "sports ball", "kite", "baseball bat",
                        "baseball glove", "skateboard", "surfboard", "tennis racket", "bottle", "wine glass", "cup",
                        "fork", "knife", "spoon", "bowl", "banana", "apple", "sandwich", "orange", "broccoli",
                        "carrot", "hot dog", "pizza", "donut", "cake", "chair", "sofa", "pottedplant", "bed",
                        "diningtable", "toilet", "tvmonitor", "laptop", "mouse", "remote", "keyboard", "cell phone",
                        "microwave", "oven", "toaster", "sink", "refrigerator", "book", "clock", "vase", "scissors",
                        "teddy bear", "hair drier", "toothbrush"
                        ]

            detected_bikes_times=[]
            detected_bikes_times_xlsx=[]
            time_count=0
            time_detect_freq=1000


            while True:

                # frame_count+=frames_detect_freq
                time_count+=time_detect_freq
                
                if time_count>1000*self.total_time:
                    break

                self.times_label.configure(text=f'{round(time_count/1000,1)}/{round(self.total_time,1)}')
                self.progressbar['value']=int(0.1*(time_count/self.total_time)) 

                self.update_status()

                cap.set(0,time_count)
                res, img = cap.read()

                results = model.predict(source=img, classes=[0,1])

                print(time_count)
                # coordinates
                for r in results:
                    boxes = r.boxes

                    for box in boxes:
                        # class name
                        cls = int(box.cls[0])

                        # bounding box
                        x1, y1, x2, y2 = box.xyxy[0]
                        x1, y1, x2, y2 = int(x1), int(y1), int(x2), int(y2) # convert to int values

                        # confidence
                        confidence = math.ceil((box.conf[0]*100))/100

                        #found_rec
                        detected_bikes_times.append(time_count)

                        detected_object=classNames[cls]

                        # save found_rec to xlsx

                        detected_bikes_times_xlsx.append((file,time_count,x1,y1,x2,y2,confidence,detected_object))


                if img is None:
                    break

                if cv2.waitKey(1) == ord('q'):
                    break

            # removing duplicated frames
            detected_bikes_times=sorted(set(detected_bikes_times))

            # saving detected_frames to xlsx

            xlsx_file='{}/times.xlsx'.format(path)
            wb=load_workbook(xlsx_file)
            ws = wb['times']
            last_empty=ws.max_row
            for i,j in enumerate(detected_bikes_times_xlsx):
                file,time_count,x1,y1,x2,y2,confidence,detected_object=j
                ws.cell(row=last_empty+1+i,column=1).value=file
                ws.cell(row=last_empty+1+i,column=2).value=time_count
                ws.cell(row=last_empty+1+i,column=3).value=x1
                ws.cell(row=last_empty+1+i,column=4).value=y1
                ws.cell(row=last_empty+1+i,column=5).value=x2
                ws.cell(row=last_empty+1+i,column=6).value=y2
                ws.cell(row=last_empty+1+i,column=7).value=confidence
                ws.cell(row=last_empty+1+i,column=8).value=detected_object

            wb.save(xlsx_file)
            wb.close()

            # update entry state

            self.entry_state[cnt+self.base_raw_count].configure(text=f'{cnt} : {file} / {'gotowe'}')
            # self.times_label.configure(text=f'{round(time_count/1000,1)}/{round(self.total_time,1)}')


            cap.release()
            cv2.destroyAllWindows()

            save_time(start_time, os.path.basename(__file__).rstrip('.py'),input_video)

    def parse_data_folder(self):
        cnt=0
        for path,_,files in os.walk('data'):
            for file in files:
                if file[-4:]=='.mp4':
                    cnt+=1
                    self.files_to_detect.append((cnt,file,path,'-'))

    def detect(self):
        # coping xlsx file
        main_dir = os.getcwd()
        raw_xlsx_file = '{}\\{}'.format(main_dir,'times.xlsx')
        date_xlsx_file = '{}\\data\\{}'.format(main_dir,'times.xlsx')
        shutil.copy(raw_xlsx_file, date_xlsx_file)

        # run script
        for cnt,file,path,_ in self.files_to_detect:
            
            self.detect_bike(cnt,file,path)

    def group(self):
        Group_frames.main()

    def cut(self):
        Cut_video.main()

    def update_status(self):
        for i in self.act_labs:
            i.update_idletasks()

if __name__ == "__main__":
    app = App()
    app.mainloop()
